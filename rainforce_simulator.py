import random
import openpyxl

tr = 0  # 표본 수
max_R = 7  # 최대 강화 수
su = 90  # 현재 강화 확률
plus = 2  # 실패 시 증가하는 확률 %
tn = 1  # 시도 횟수
rf = []  # 강화 시뮬레이터 배열
pxl = openpyxl.Workbook()  # 저장할 엑셀 파일 지정
sheet = pxl.active
R = 1  # 엑셀 상수
C = 1  # 엑셀 상수
name = ''  # 엑셀 파일 이름


def rain_force_C(nc):  # 해당 강화 확률로 초기화시켜주는 함수
    global su, rf
    su = 90
    rf = [0 for i in range(1000)]
    for z in range(1, nc):
        if z < 5:
            su -= 15
        elif 5 <= z < 6:
            su = 10
        elif z == 6:
            su = 1
    sun = su * 10  # 강화 확률 상수
    while sun != 0:
        ran = random.randrange(0, 1000)
        if rf[ran] == 0:  # 강화 실패 숫자는 0, 성공 숫자는 1
            rf[ran] = 1
            sun -= 1


def plus_p():
    global rf
    plus_r = plus * 10  # 강화 확률 증가 상수
    while plus_r != 0:
        ran = random.randrange(0, 1000)
        if rf[ran] == 0:
            rf[ran] = 1
            plus_r -= 1


def rain_force_a_z(tri):  # 0부터 7까지 강화시켜주는 함수
    global max_R, rf, tn, R, C, sheet, name, pxl
    openpyxl.load_workbook(name + ".xlsx")
    sheet = pxl.active
    rn = 0  # 현재 강화 수
    t1 = 3  # 첫 시도 성공 횟수로 넘어가는 상수
    t2 = 2  # 성공 횟수로 넘어가는 상수
    t3 = 1  # 실패 횟수로 넘어가는 상수
    R += 1
    sheet.cell(row=R, column=C).value = '1 - 7강'
    R += 1
    sheet.cell(row=R, column=C).value = '목표 단계'
    for i in range(2, 9):
        sheet.cell(row=R, column=i).value = i - 1
    R += 1
    sheet.cell(row=R, column=C).value = '원트횟수'
    for i in range(2, 9):
        sheet.cell(row=R, column=i).value = 0
    R += 1
    sheet.cell(row=R, column=C).value = '성공횟수'
    for i in range(2, 9):
        sheet.cell(row=R, column=i).value = 0
    R += 1
    sheet.cell(row=R, column=C).value = '실패횟수'
    for i in range(2, 9):
        sheet.cell(row=R, column=i).value = 0
    R += 1
    rain_force_C(rn + 1)
    while tri != 0:
        res = rf[random.randrange(0, 1000)]
        if res == 0:
            sheet.cell(row=R - t3, column=C + 1 + rn).value += 1
            if rn < 5:
                plus_p()
            tn += 1
        elif res == 1:
            sheet.cell(row=R - t2, column=C + 1 + rn).value += 1
            if tn == 1:
                sheet.cell(row=R - t1, column=C + 1 + rn).value += 1
            if rn == max_R - 1:
                rn = 0
            else:
                rn += 1
            rain_force_C(rn + 1)
            tn = 1
        else:
            print('오류발생')
            exit(0)
        tri -= 1
    pxl.save(name + '.xlsx')
    retry()


def rain_force_o(tri, n):  # 한 강화단계만 강화시켜주는 함수
    global rf, tn, R, C, sheet, name, pxl
    openpyxl.load_workbook(name + ".xlsx")
    sheet = pxl.active
    rain_force_C(n + 1)
    tn = 1
    t1 = 3  # 첫 시도 성공 횟수로 넘어가는 상수
    t2 = 2  # 성공 횟수로 넘어가는 상수
    t3 = 1  # 실패 횟수로 넘어가는 상수
    R += 1
    text = str(n+1)
    sheet.cell(row=R, column=C).value = text + '강'
    R += 1
    sheet.cell(row=R, column=C).value = '목표 단계'
    sheet.cell(row=R, column=C + 1).value = n + 1
    R += 1
    sheet.cell(row=R, column=C).value = '원트횟수'
    sheet.cell(row=R, column=C + 1).value = 0
    R += 1
    sheet.cell(row=R, column=C).value = '성공횟수'
    sheet.cell(row=R, column=C + 1).value = 0
    R += 1
    sheet.cell(row=R, column=C).value = '실패횟수'
    sheet.cell(row=R, column=C + 1).value = 0
    R += 1
    while tri != 0:
        res = rf[random.randrange(0, 1000)]
        if res == 0:
            sheet.cell(row=R - t3, column=C + 1).value += 1
            if n < 5:
                plus_p()
            tn += 1
        elif res == 1:
            sheet.cell(row=R - t2, column=C + 1).value += 1
            if tn == 1:
                sheet.cell(row=R - t1, column=C + 1).value += 1
            tn = 1
            rain_force_C(n + 1)
        else:
            print('오류발생')
            exit(0)
        tri -= 1
    pxl.save(name + '.xlsx')
    retry()


def retry():  # 표본 수집 후 나오는 화면
    re = input('계속 하시겠습니까(y/n):')
    if re == 'y':
        return menu()
    elif re == 'n':
        return 0
    else:
        print('잘못 입력하셨습니다.')
        return retry()


def menu():  # 메인메뉴
    global R, C, name, pxl, sheet
    how = int(input('표본 수집 방식을 입력하세요(1-7까지는 1번, 같은 강화단계만은 2번):'))
    if how == 1:
        rain_force_a_z(tr)
    elif how == 2:
        menu_2()
    else:
        print('잘못 입력하셨습니다.')
        return menu()


def menu_2():  # 강화 단계 선택 메뉴
    global tr
    n = int(input('필요한 표본의 현재 강화 단계를 입력해주세요:'))
    if n > 6:
        if n == 7:
            print('최대 강화 단계는 강화할 수 없습니다.')
            return menu_2()
        else:
            print('잘못된 값을 입력하셨습니다.')
            return menu_2()
    elif 0 <= n <= 6:
        rain_force_o(tr, n)
    else:
        print('잘못된 값을 입력하셨습니다.')
        return menu_2()


def sample():
    global pxl, sheet, name, R, C, tr
    pxl = openpyxl.load_workbook(name + ".xlsx")
    sheet = pxl.active
    tr = int(input('강화에 필요한 표본 수를 입력하세요:'))
    text = str(tr)
    sheet.cell(row=R, column=C).value = '표본수: ' + text
    R += 1
    pxl.save(name + '.xlsx')


def save_P():  # 엑셀 파일 이름 저장
    global name
    name = input('저장될 파일 이름을 입력하세요:')
    pxl.save(name + '.xlsx')


save_P()
sample()
menu()
